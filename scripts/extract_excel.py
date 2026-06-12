from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_EXCEL = Path(r"C:\Users\meche\Downloads\celestino2.xlsx")
DEFAULT_PLAYERS = [
    {
        "id": "miguel",
        "name": "Miguel",
        "excel": DEFAULT_EXCEL,
    },
    {
        "id": "gonzalo",
        "name": "Gonzalo",
        "excel": Path(r"C:\Users\meche\Downloads\celestino Echevarria_mundial-2026.xlsx"),
    },
]
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "data" / "predictions.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def stage_for(match_no: int) -> str:
    if match_no <= 72:
        return "Grupos"
    if match_no <= 88:
        return "Dieciseisavos"
    if match_no <= 96:
        return "Octavos"
    if match_no <= 100:
        return "Cuartos"
    if match_no <= 102:
        return "Semifinales"
    if match_no == 103:
        return "Tercer puesto"
    if match_no == 104:
        return "Final"
    return "Otro"


def as_int(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def read_teams(ws):
    flags = {}
    for row in range(2, 50):
        name = ws.cell(row, 11).value
        flag = ws.cell(row, 12).value
        if name and flag and not str(flag).startswith("#"):
            flags[str(name)] = str(flag)

    teams = []
    for row in range(2, 50):
        number = as_int(ws.cell(row, 1).value)
        name = ws.cell(row, 2).value
        group = ws.cell(row, 3).value
        rank = as_int(ws.cell(row, 4).value)
        if not number or not name or not group:
            continue
        teams.append(
            {
                "id": number,
                "name": str(name),
                "group": str(group),
                "rank": rank,
                "flag": flags.get(str(name), ""),
            }
        )
    return teams


def read_matches(ws):
    matches = []
    current_group = None
    for row in range(1, ws.max_row + 1):
        maybe_group = ws.cell(row, 2).value
        if isinstance(maybe_group, str) and len(maybe_group) == 1 and "A" <= maybe_group <= "L":
            current_group = maybe_group

        match_no = as_int(ws.cell(row, 34).value)
        home = ws.cell(row, 27).value
        away = ws.cell(row, 32).value
        home_goals = as_int(ws.cell(row, 29).value)
        away_goals = as_int(ws.cell(row, 30).value)

        if not match_no or not home or not away:
            continue

        match_group = current_group if match_no <= 72 else None
        matches.append(
            {
                "id": match_no,
                "excelRow": row,
                "stage": stage_for(match_no),
                "group": match_group,
                "dateTime": clean(ws.cell(row, 24).value),
                "matchday": clean(ws.cell(row, 26).value),
                "slotHome": clean(ws.cell(row, 1).value),
                "slotAway": clean(ws.cell(row, 2).value),
                "homeTeam": str(home),
                "awayTeam": str(away),
                "homeGoals": home_goals,
                "awayGoals": away_goals,
            }
        )

    return sorted(matches, key=lambda item: item["id"])


def read_honors(ws):
    return {
        "champion": clean(ws.cell(150, 27).value),
        "runnerUp": clean(ws.cell(151, 27).value),
        "third": clean(ws.cell(152, 27).value),
        "topScorers": [clean(ws.cell(row, 27).value) for row in (154, 155, 156) if ws.cell(row, 27).value],
        "bestPlayers": [clean(ws.cell(row, 27).value) for row in (158, 159, 160) if ws.cell(row, 27).value],
    }


def slugify(value):
    text = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return text or "participante"


def read_prediction(excel_path: Path, player_id: str, player_name: str):
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    worldcup = workbook["WORLDCUP"]
    teams_sheet = workbook["Equipos"]

    return {
        "id": player_id,
        "name": player_name,
        "tournament": clean(worldcup["A1"].value) or "XXIII Mundial USA/MEX/CAN 2026",
        "sourceFile": str(excel_path),
        "teams": read_teams(teams_sheet),
        "matches": read_matches(worldcup),
        "honors": read_honors(worldcup),
    }


def parse_player_arg(value: str):
    parts = value.split(":", 2)
    if len(parts) != 3:
        raise argparse.ArgumentTypeError("Usa el formato id:nombre:ruta_excel")
    player_id, player_name, excel = parts
    return {
        "id": slugify(player_id),
        "name": player_name.strip() or player_id.strip(),
        "excel": Path(excel),
    }


def player_inputs(args):
    if args.player:
        return args.player

    if args.excel:
        excel = Path(args.excel)
        name = args.name or excel.stem
        return [
            {
                "id": slugify(args.id or name),
                "name": name,
                "excel": excel,
            }
        ]

    return DEFAULT_PLAYERS


def main():
    parser = argparse.ArgumentParser(description="Extrae la porra del Excel del Mundial 2026 a JSON.")
    parser.add_argument("excel", nargs="?")
    parser.add_argument("--id", default=None, help="Id del participante si se extrae un solo Excel.")
    parser.add_argument("--name", default=None, help="Nombre del participante si se extrae un solo Excel.")
    parser.add_argument(
        "--player",
        action="append",
        type=parse_player_arg,
        help="Participante en formato id:nombre:ruta_excel. Se puede repetir.",
    )
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    output_path = Path(args.output)
    participants = []

    for player in player_inputs(args):
        excel_path = player["excel"]
        if not excel_path.exists():
            raise FileNotFoundError(f"No existe el Excel de {player['name']}: {excel_path}")
        prediction = read_prediction(excel_path, player["id"], player["name"])
        participants.append(prediction)

    if not participants:
        raise RuntimeError("No se encontro ninguna porra para exportar.")

    primary = participants[0]
    payload = {
        "tournament": primary["tournament"],
        "sourceFile": primary["sourceFile"],
        "sourceFiles": {player["id"]: player["sourceFile"] for player in participants},
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "participants": participants,
        "teams": primary["teams"],
        "matches": primary["matches"],
        "honors": primary["honors"],
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    names = ", ".join(player["name"] for player in participants)
    print(
        f"Escrito {output_path} con {len(participants)} porras ({names}), "
        f"{len(primary['teams'])} equipos y {len(primary['matches'])} partidos por porra."
    )


if __name__ == "__main__":
    main()
